#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube 브랜드 분석 도구 - 테이블 설명서 생성기
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def create_documentation():
    """테이블 설명서 엑셀 파일 생성"""
    
    # Videos 테이블 설명
    videos_data = [
        # [컬럼명, 영어명, 한국어명, 설명, 데이터타입, API소스]
        ["video_id", "Video ID", "비디오 ID", "YouTube 비디오의 고유 식별자 (11자리)", "string", "videos.id"],
        ["search_keyword", "Search Keyword", "검색 키워드", "비디오 검색에 사용된 키워드", "string", "custom"],
        ["search_region", "Search Region", "검색 지역", "비디오 검색 대상 지역 코드 (US, KR 등)", "string", "search.regionCode"],
        ["search_order", "Search Order", "정렬 순서", "검색 결과 정렬 방식 (relevance, date 등)", "string", "search.order"],
        ["collected_at", "Collected At", "수집 시각", "데이터가 수집된 시각 (ISO 8601)", "datetime", "custom"],
        
        # snippet 정보
        ["title", "Title", "제목", "비디오 제목", "string", "snippet.title"],
        ["description", "Description", "설명", "비디오 설명 (개행문자 제거됨)", "string", "snippet.description"],
        ["channel_id", "Channel ID", "채널 ID", "비디오를 업로드한 채널의 고유 ID", "string", "snippet.channelId"],
        ["channel_title", "Channel Title", "채널명", "채널 이름", "string", "snippet.channelTitle"],
        ["published_at", "Published At", "게시 시각", "비디오가 YouTube에 업로드된 시각", "datetime", "snippet.publishedAt"],
        ["category_id", "Category ID", "카테고리 ID", "YouTube 카테고리 ID (28=과학기술, 24=엔터테인먼트 등)", "string", "snippet.categoryId"],
        ["default_language", "Default Language", "기본 언어", "비디오의 기본 언어 코드", "string", "snippet.defaultLanguage"],
        ["default_audio_language", "Default Audio Language", "기본 오디오 언어", "비디오 오디오의 기본 언어", "string", "snippet.defaultAudioLanguage"],
        ["live_broadcast_content", "Live Broadcast Content", "라이브 방송 콘텐츠", "라이브 방송 여부 (none, live, upcoming)", "string", "snippet.liveBroadcastContent"],
        ["tags", "Tags", "태그", "비디오 태그들 (콤마로 구분)", "string", "snippet.tags"],
        
        # 썸네일 정보
        ["thumbnail_default", "Thumbnail Default", "기본 썸네일", "기본 크기 썸네일 URL (120x90)", "string", "snippet.thumbnails.default"],
        ["thumbnail_medium", "Thumbnail Medium", "중간 썸네일", "중간 크기 썸네일 URL (320x180)", "string", "snippet.thumbnails.medium"],
        ["thumbnail_high", "Thumbnail High", "고화질 썸네일", "고화질 썸네일 URL (480x360)", "string", "snippet.thumbnails.high"],
        ["thumbnail_standard", "Thumbnail Standard", "표준 썸네일", "표준 썸네일 URL (640x480)", "string", "snippet.thumbnails.standard"],
        ["thumbnail_maxres", "Thumbnail Max Resolution", "최고화질 썸네일", "최고화질 썸네일 URL (1280x720)", "string", "snippet.thumbnails.maxres"],
        
        # 통계 정보
        ["view_count", "View Count", "조회수", "비디오 총 조회수", "integer", "statistics.viewCount"],
        ["like_count", "Like Count", "좋아요 수", "비디오 좋아요 수", "integer", "statistics.likeCount"],
        ["dislike_count", "Dislike Count", "싫어요 수", "비디오 싫어요 수 (현재 비공개)", "integer", "statistics.dislikeCount"],
        ["favorite_count", "Favorite Count", "즐겨찾기 수", "즐겨찾기에 추가된 횟수", "integer", "statistics.favoriteCount"],
        ["comment_count", "Comment Count", "댓글 수", "비디오 댓글 총 개수", "integer", "statistics.commentCount"],
        
        # 콘텐츠 상세 정보
        ["duration_seconds", "Duration Seconds", "길이(초)", "비디오 길이를 초 단위로 변환", "integer", "contentDetails.duration"],
        ["duration_iso", "Duration ISO", "길이(ISO)", "비디오 길이 ISO 8601 형식 (PT4M13S)", "string", "contentDetails.duration"],
        ["dimension", "Dimension", "화면 차원", "2d 또는 3d", "string", "contentDetails.dimension"],
        ["definition", "Definition", "화질", "sd (표준화질) 또는 hd (고화질)", "string", "contentDetails.definition"],
        ["caption", "Caption", "자막", "자막 사용 가능 여부 (true/false)", "string", "contentDetails.caption"],
        ["licensed_content", "Licensed Content", "라이선스 콘텐츠", "라이선스된 콘텐츠 여부", "boolean", "contentDetails.licensedContent"],
        ["content_rating", "Content Rating", "콘텐츠 등급", "콘텐츠 등급 정보", "string", "contentDetails.contentRating"],
        ["projection", "Projection", "프로젝션", "비디오 프로젝션 타입 (rectangular, 360)", "string", "contentDetails.projection"],
        ["has_custom_thumbnail", "Has Custom Thumbnail", "커스텀 썸네일 여부", "사용자 지정 썸네일 사용 여부", "boolean", "contentDetails.hasCustomThumbnail"],
        
        # 상태 정보
        ["upload_status", "Upload Status", "업로드 상태", "업로드 처리 상태 (processed, uploaded 등)", "string", "status.uploadStatus"],
        ["privacy_status", "Privacy Status", "공개 설정", "비디오 공개 설정 (public, private, unlisted)", "string", "status.privacyStatus"],
        ["license", "License", "라이선스", "비디오 라이선스 (youtube, creativeCommon)", "string", "status.license"],
        ["embeddable", "Embeddable", "임베드 가능", "다른 사이트 임베드 허용 여부", "boolean", "status.embeddable"],
        ["public_stats_viewable", "Public Stats Viewable", "공개 통계 표시", "공개 통계 표시 여부", "boolean", "status.publicStatsViewable"],
        ["made_for_kids", "Made For Kids", "어린이 대상", "어린이 대상 콘텐츠 여부", "boolean", "status.madeForKids"],
        ["self_declared_made_for_kids", "Self Declared Made For Kids", "자체 신고 어린이 대상", "제작자가 스스로 어린이 대상으로 신고", "boolean", "status.selfDeclaredMadeForKids"],
        
        # 주제 정보
        ["topic_ids", "Topic IDs", "주제 ID", "위키피디아 주제 ID들 (콤마 구분)", "string", "topicDetails.topicIds"],
        ["relevant_topic_ids", "Relevant Topic IDs", "관련 주제 ID", "관련 위키피디아 주제 ID들", "string", "topicDetails.relevantTopicIds"],
        ["topic_categories", "Topic Categories", "주제 카테고리", "위키피디아 주제 카테고리 URL", "string", "topicDetails.topicCategories"],
        
        # 녹화 정보
        ["recording_date", "Recording Date", "녹화 날짜", "비디오가 녹화된 날짜", "datetime", "recordingDetails.recordingDate"],
        ["location_latitude", "Location Latitude", "위치 위도", "녹화 위치의 위도", "float", "recordingDetails.location.latitude"],
        ["location_longitude", "Location Longitude", "위치 경도", "녹화 위치의 경도", "float", "recordingDetails.location.longitude"],
        ["location_altitude", "Location Altitude", "위치 고도", "녹화 위치의 고도", "float", "recordingDetails.location.altitude"],
        
        # 채널 정보 (병합)
        ["channel_subscriber_count", "Channel Subscriber Count", "채널 구독자 수", "채널의 총 구독자 수", "integer", "channels.statistics.subscriberCount"],
        ["channel_video_count", "Channel Video Count", "채널 비디오 수", "채널의 총 비디오 개수", "integer", "channels.statistics.videoCount"],
        ["channel_total_view_count", "Channel Total View Count", "채널 총 조회수", "채널의 총 조회수", "integer", "channels.statistics.viewCount"],
        ["channel_description", "Channel Description", "채널 설명", "채널 설명 (개행문자 제거됨)", "string", "channels.snippet.description"],
        ["channel_country", "Channel Country", "채널 국가", "채널이 등록된 국가 코드", "string", "channels.snippet.country"],
        ["channel_custom_url", "Channel Custom URL", "채널 커스텀 URL", "채널의 사용자 지정 URL", "string", "channels.snippet.customUrl"],
        ["channel_published_at", "Channel Published At", "채널 개설일", "채널이 생성된 날짜", "datetime", "channels.snippet.publishedAt"],
    ]
    
    # Comments 테이블 설명
    comments_data = [
        # [컬럼명, 영어명, 한국어명, 설명, 데이터타입, API소스]
        ["video_id", "Video ID", "비디오 ID", "댓글이 달린 비디오의 ID", "string", "commentThreads.snippet.videoId"],
        ["comment_id", "Comment ID", "댓글 ID", "댓글의 고유 식별자", "string", "commentThreads.snippet.topLevelComment.id"],
        ["comment_type", "Comment Type", "댓글 타입", "최상위 댓글(top_level) 또는 답글(reply)", "string", "custom"],
        ["parent_comment_id", "Parent Comment ID", "부모 댓글 ID", "답글인 경우 상위 댓글의 ID", "string", "custom"],
        ["collected_at", "Collected At", "수집 시각", "데이터가 수집된 시각", "datetime", "custom"],
        
        # 작성자 정보
        ["author_display_name", "Author Display Name", "작성자 표시명", "댓글 작성자의 표시 이름", "string", "snippet.authorDisplayName"],
        ["author_profile_image_url", "Author Profile Image URL", "작성자 프로필 이미지 URL", "작성자 프로필 이미지 URL", "string", "snippet.authorProfileImageUrl"],
        ["author_channel_url", "Author Channel URL", "작성자 채널 URL", "작성자 채널 페이지 URL", "string", "snippet.authorChannelUrl"],
        ["author_channel_id", "Author Channel ID", "작성자 채널 ID", "작성자 채널의 고유 ID", "string", "snippet.authorChannelId.value"],
        
        # 댓글 내용
        ["comment_text_display", "Comment Text Display", "댓글 표시 텍스트", "표시용 댓글 텍스트 (HTML 태그 포함)", "string", "snippet.textDisplay"],
        ["comment_text_original", "Comment Text Original", "댓글 원본 텍스트", "원본 댓글 텍스트 (플레인 텍스트)", "string", "snippet.textOriginal"],
        ["comment_text_length", "Comment Text Length", "댓글 길이", "댓글 텍스트의 문자 수", "integer", "custom"],
        
        # 상호작용 정보
        ["like_count", "Like Count", "좋아요 수", "댓글에 달린 좋아요 수", "integer", "snippet.likeCount"],
        ["reply_count", "Reply Count", "답글 수", "댓글에 달린 답글 개수", "integer", "snippet.totalReplyCount"],
        ["moderation_status", "Moderation Status", "검토 상태", "댓글 검토 상태", "string", "snippet.moderationStatus"],
        
        # 시간 정보
        ["published_at", "Published At", "게시 시각", "댓글이 작성된 시각", "datetime", "snippet.publishedAt"],
        ["updated_at", "Updated At", "수정 시각", "댓글이 마지막으로 수정된 시각", "datetime", "snippet.updatedAt"],
        
        # 추가 메타데이터
        ["viewer_rating", "Viewer Rating", "시청자 평가", "현재 사용자의 댓글 평가", "string", "snippet.viewerRating"],
        ["can_rate", "Can Rate", "평가 가능", "댓글 평가 가능 여부", "boolean", "snippet.canRate"],
        
        # 감정 분석 결과 (추가 분석)
        ["sentiment_score", "Sentiment Score", "감정 점수", "감정 분석 점수 (-1=부정, 0=중립, 1=긍정)", "float", "custom_analysis"],
        ["sentiment_category", "Sentiment Category", "감정 카테고리", "감정 분류 (positive, negative, neutral)", "string", "custom_analysis"],
        ["subjectivity_score", "Subjectivity Score", "주관성 점수", "주관성 점수 (0=객관적, 1=주관적)", "float", "custom_analysis"],
        ["brand_mentions", "Brand Mentions", "브랜드 언급", "댓글에 언급된 브랜드들", "string", "custom_analysis"],
        ["competitor_mentions", "Competitor Mentions", "경쟁사 언급", "댓글에 언급된 경쟁사들", "string", "custom_analysis"],
        ["analyzed_at", "Analyzed At", "분석 시각", "감정 분석이 수행된 시각", "datetime", "custom_analysis"],
    ]
    
    # DataFrame 생성
    videos_df = pd.DataFrame(videos_data, columns=['컬럼명', '영어명', '한국어명', '설명', '데이터타입', 'API소스'])
    comments_df = pd.DataFrame(comments_data, columns=['컬럼명', '영어명', '한국어명', '설명', '데이터타입', 'API소스'])
    
    # 엑셀 파일 생성
    with pd.ExcelWriter('data/YouTube_API_Table_Documentation.xlsx', engine='openpyxl') as writer:
        # Videos 시트
        videos_df.to_excel(writer, sheet_name='Videos Table', index=False)
        
        # Comments 시트
        comments_df.to_excel(writer, sheet_name='Comments Table', index=False)
        
        # 개요 시트
        overview_data = [
            ['테이블명', '설명', '레코드수', '컬럼수', '주요용도'],
            ['Videos', 'YouTube 비디오의 모든 메타데이터와 통계 정보', '100개', '55개', '비디오 분석, 채널 분석, 성과 측정'],
            ['Comments', 'YouTube 댓글과 감정분석 결과', '1,000+개', '25개', '감정 분석, 브랜드 언급, 사용자 반응 분석'],
        ]
        overview_df = pd.DataFrame(overview_data[1:], columns=overview_data[0])
        overview_df.to_excel(writer, sheet_name='Overview', index=False)
        
        # API 정보 시트
        api_info_data = [
            ['항목', '내용'],
            ['API 버전', 'YouTube Data API v3'],
            ['수집 방식', 'search.list → videos.list → commentThreads.list'],
            ['페이지네이션', '50개씩 2회 호출로 100개 수집'],
            ['API 할당량', '36 units (일일 10,000 units 제한)'],
            ['지역 필터', 'US (미국 지역)'],
            ['시간 필터', '최근 90일 내 업로드'],
            ['정렬 기준', 'relevance (관련성 높은 순)'],
            ['댓글 수집', '상위 30개 비디오만 (API 절약)'],
            ['감정 분석', 'TextBlob + 키워드 기반 분석'],
        ]
        api_info_df = pd.DataFrame(api_info_data[1:], columns=api_info_data[0])
        api_info_df.to_excel(writer, sheet_name='API Info', index=False)
    
    # 스타일 적용을 위해 workbook 다시 열기
    from openpyxl import load_workbook
    wb = load_workbook('data/YouTube_API_Table_Documentation.xlsx')
    
    # 각 시트에 스타일 적용
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 헤더 스타일 (첫 번째 행)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    wb.save('data/YouTube_API_Table_Documentation.xlsx')
    
    print("YouTube API 테이블 설명서가 생성되었습니다!")
    print("파일 위치: data/YouTube_API_Table_Documentation.xlsx")
    print("포함된 시트:")
    print("   - Overview: 전체 개요")
    print("   - Videos Table: 비디오 테이블 필드 설명")
    print("   - Comments Table: 댓글 테이블 필드 설명")
    print("   - API Info: API 수집 정보")

if __name__ == "__main__":
    create_documentation()